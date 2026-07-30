"""
Generate ALFAGRAND Product Data Excel Template
One file, multiple sheets — one sheet per category, one row per model variant
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT = "C:/Users/华为/.qclaw/workspace-37i6raipm851ul5j/ALFAGRAND_Product_Data_Template.xlsx"

# ── Color palette ──────────────────────────────────────────────
C_TITLE_BG  = "1A2535"   # dark header bg
C_TITLE_FG  = "00E5FF"   # cyan text
C_HEAD_BG   = "0D3047"   # column header bg
C_HEAD_FG   = "FFFFFF"   # white text
C_SUB_BG    = "112233"   # series info row bg
C_SUB_FG    = "E0F7FF"
C_ROW_A     = "1C2E40"   # alternating row A
C_ROW_B     = "172437"   # alternating row B
C_INPUT_BG  = "FFF9E6"   # yellow — cells user must fill
C_HINT_FG   = "888888"   # grey hint text
C_BORDER    = "2A4560"
C_NOTE_BG   = "0A3D62"
C_NOTE_FG   = "AADDFF"
C_IMG_BG    = "1A3A1A"   # green tint — image cells
C_IMG_FG    = "88FF88"

def thin_border():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, value, bg=C_HEAD_BG, fg=C_HEAD_FG, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=9)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=wrap)
    c.border = thin_border()
    return c

def data_cell(ws, row, col, value="", bg=C_ROW_A, fg="D0E8FF",
              bold=False, wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", color=fg, size=9, bold=bold, italic=italic)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=wrap)
    c.border = thin_border()
    return c

def input_cell(ws, row, col, value="", hint=""):
    """Yellow 'fill me in' cell"""
    c = ws.cell(row=row, column=col, value=value if value else hint)
    c.font = Font(name="Arial", color="111111" if value else C_HINT_FG,
                  size=9, italic=not bool(value))
    c.fill = PatternFill("solid", fgColor=C_INPUT_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border()
    return c

def image_cell(ws, row, col, value=""):
    """Green tint — image filename cell"""
    hint = "(e.g. my-pump.jpg)"
    c = ws.cell(row=row, column=col, value=value if value else hint)
    c.font = Font(name="Arial", color=C_IMG_FG if value else C_HINT_FG,
                  size=9, italic=not bool(value))
    c.fill = PatternFill("solid", fgColor=C_IMG_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border()
    return c


# ══════════════════════════════════════════════════════════════
# DATA — existing series (current placeholder data, marked so user knows what to replace)
# ══════════════════════════════════════════════════════════════

CATEGORIES = [
    {
        "slug": "peripheral-pump",
        "name": "Peripheral Pump（漩涡泵）",
        "series": [
            {
                "series_key": "pm-series",
                "series_name": "PM SERIES",
                "series_subtitle": "High-Pressure Peripheral Pumps",
                "image_hero": "/images/products/peripheral-pump.png",
                "image_overview": "/images/products/peripheral-pump.png",
                "models": [
                    {"model":"PM45A","powerHP":"0.37","powerKW":"0.25","maxHead":"30","maxFlow":"1.2 m³/h","suctionHead":"7","inletOutlet":"¾\" x ¾\"","weight":"8.5","image":""},
                    {"model":"PM45B","powerHP":"0.5","powerKW":"0.37","maxHead":"35","maxFlow":"1.5 m³/h","suctionHead":"7","inletOutlet":"¾\" x ¾\"","weight":"10.2","image":""},
                    {"model":"PM45C","powerHP":"0.75","powerKW":"0.55","maxHead":"40","maxFlow":"1.8 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"12.8","image":""},
                    {"model":"PM45D","powerHP":"1.0","powerKW":"0.75","maxHead":"45","maxFlow":"2.1 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"15.5","image":""},
                    {"model":"PM60A","powerHP":"0.5","powerKW":"0.37","maxHead":"40","maxFlow":"1.8 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"12.5","image":""},
                    {"model":"PM60B","powerHP":"0.75","powerKW":"0.55","maxHead":"50","maxFlow":"2.4 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"15.8","image":""},
                    {"model":"PM60C","powerHP":"1.0","powerKW":"0.75","maxHead":"55","maxFlow":"2.7 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"18.2","image":""},
                    {"model":"PM60D","powerHP":"1.5","powerKW":"1.1","maxHead":"60","maxFlow":"3.0 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"21.5","image":""},
                    {"model":"PM80A","powerHP":"1.5","powerKW":"1.1","maxHead":"65","maxFlow":"2.7 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"22.0","image":""},
                    {"model":"PM80B","powerHP":"2.0","powerKW":"1.5","maxHead":"72","maxFlow":"3.3 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"25.5","image":""},
                    {"model":"PM80C","powerHP":"2.5","powerKW":"1.85","maxHead":"80","maxFlow":"3.9 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"29.0","image":""},
                ],
            },
            {
                "series_key": "qb-series",
                "series_name": "QB SERIES",
                "series_subtitle": "Standard Peripheral Pumps",
                "image_hero": "/images/products/peripheral-pump.png",
                "image_overview": "/images/products/peripheral-pump.png",
                "models": [
                    {"model":"QB60","powerHP":"0.5","powerKW":"0.37","maxHead":"28","maxFlow":"1.2 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"8.5","image":""},
                    {"model":"QB70","powerHP":"0.75","powerKW":"0.55","maxHead":"35","maxFlow":"1.5 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"10.5","image":""},
                    {"model":"QB80","powerHP":"1.0","powerKW":"0.75","maxHead":"42","maxFlow":"1.8 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"12.5","image":""},
                    {"model":"QB90","powerHP":"1.5","powerKW":"1.1","maxHead":"50","maxFlow":"2.4 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"15.0","image":""},
                ],
            },
            {
                "series_key": "wzb-series",
                "series_name": "WZB SERIES",
                "series_subtitle": "Self-Priming Peripheral Pumps",
                "image_hero": "/images/products/peripheral-pump.png",
                "image_overview": "/images/products/peripheral-pump.png",
                "models": [
                    {"model":"WZB250","powerHP":"0.33","powerKW":"0.25","maxHead":"25","maxFlow":"0.9 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"7.8","image":""},
                    {"model":"WZB370","powerHP":"0.5","powerKW":"0.37","maxHead":"32","maxFlow":"1.2 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"9.2","image":""},
                    {"model":"WZB550","powerHP":"0.75","powerKW":"0.55","maxHead":"40","maxFlow":"1.5 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"11.5","image":""},
                ],
            },
            {
                "series_key": "awzb-series",
                "series_name": "AWZB SERIES",
                "series_subtitle": "Automatic Self-Priming Peripheral Pumps",
                "image_hero": "/images/products/peripheral-pump.png",
                "image_overview": "/images/products/peripheral-pump.png",
                "models": [
                    {"model":"AWZB250","powerHP":"0.33","powerKW":"0.25","maxHead":"25","maxFlow":"0.9 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"8.0","image":""},
                    {"model":"AWZB370","powerHP":"0.5","powerKW":"0.37","maxHead":"32","maxFlow":"1.2 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"9.8","image":""},
                    {"model":"AWZB550","powerHP":"0.75","powerKW":"0.55","maxHead":"40","maxFlow":"1.5 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"12.0","image":""},
                ],
            },
            {
                "series_key": "mkp-series",
                "series_name": "MKP SERIES",
                "series_subtitle": "Garden & Domestic Peripheral Pumps",
                "image_hero": "/images/products/peripheral-pump.png",
                "image_overview": "/images/products/peripheral-pump.png",
                "models": [
                    {"model":"MKP250","powerHP":"0.33","powerKW":"0.25","maxHead":"22","maxFlow":"0.8 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"5.5","image":""},
                    {"model":"MKP370","powerHP":"0.5","powerKW":"0.37","maxHead":"28","maxFlow":"1.0 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"6.8","image":""},
                ],
            },
        ],
    },
    {
        "slug": "centrifugal-pump",
        "name": "Centrifugal Pump（离心泵）",
        "series": [
            {
                "series_key": "cnpm-series",
                "series_name": "CNPm SERIES",
                "series_subtitle": "Monobloc Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"CNPm50","powerHP":"0.5","powerKW":"0.37","maxHead":"32","maxFlow":"50 L/min","suctionHead":"7","inletOutlet":"1.5\" x 1.5\"","weight":"7.2","image":""},
                    {"model":"CNPm75","powerHP":"0.75","powerKW":"0.55","maxHead":"45","maxFlow":"80 L/min","suctionHead":"7","inletOutlet":"1.5\" x 1.5\"","weight":"10.5","image":""},
                    {"model":"CNPm100","powerHP":"1.0","powerKW":"0.75","maxHead":"48","maxFlow":"110 L/min","suctionHead":"7","inletOutlet":"1.5\" x 1.5\"","weight":"13.0","image":""},
                    {"model":"CNPm150","powerHP":"1.5","powerKW":"1.1","maxHead":"52","maxFlow":"140 L/min","suctionHead":"7","inletOutlet":"2\" x 2\"","weight":"16.5","image":""},
                    {"model":"CNPm200","powerHP":"2.0","powerKW":"1.5","maxHead":"58","maxFlow":"170 L/min","suctionHead":"7","inletOutlet":"2\" x 2\"","weight":"20.0","image":""},
                    {"model":"CNPm300","powerHP":"3.0","powerKW":"2.2","maxHead":"62","maxFlow":"200 L/min","suctionHead":"7","inletOutlet":"2\" x 2\"","weight":"25.0","image":""},
                ],
            },
            {
                "series_key": "cnpl-series",
                "series_name": "CNPl SERIES",
                "series_subtitle": "Large-Flow Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"CNPl300","powerHP":"3.0","powerKW":"2.2","maxHead":"28","maxFlow":"400 L/min","suctionHead":"6","inletOutlet":"3\" x 3\"","weight":"35.0","image":""},
                    {"model":"CNPl550","powerHP":"5.0","powerKW":"4.0","maxHead":"32","maxFlow":"700 L/min","suctionHead":"6","inletOutlet":"4\" x 4\"","weight":"52.0","image":""},
                    {"model":"CNPl750","powerHP":"7.5","powerKW":"5.5","maxHead":"35","maxFlow":"1000 L/min","suctionHead":"6","inletOutlet":"4\" x 4\"","weight":"68.0","image":""},
                ],
            },
            {
                "series_key": "cpm-series",
                "series_name": "CPm SERIES",
                "series_subtitle": "High-Efficiency Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"CPm130","powerHP":"0.5","powerKW":"0.37","maxHead":"30","maxFlow":"55 L/min","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"6.5","image":""},
                    {"model":"CPm158","powerHP":"0.75","powerKW":"0.55","maxHead":"38","maxFlow":"80 L/min","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"7.8","image":""},
                    {"model":"CPm600","powerHP":"1.0","powerKW":"0.75","maxHead":"44","maxFlow":"100 L/min","suctionHead":"7","inletOutlet":"1.25\" x 1\"","weight":"9.5","image":""},
                ],
            },
            {
                "series_key": "mcp-series",
                "series_name": "MCP SERIES",
                "series_subtitle": "Multi-Purpose Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"MCP100","powerHP":"1.0","powerKW":"0.75","maxHead":"40","maxFlow":"120 L/min","suctionHead":"7","inletOutlet":"1.5\" x 1.5\"","weight":"14.0","image":""},
                    {"model":"MCP150","powerHP":"1.5","powerKW":"1.1","maxHead":"48","maxFlow":"180 L/min","suctionHead":"7","inletOutlet":"2\" x 2\"","weight":"18.5","image":""},
                ],
            },
            {
                "series_key": "dk-series",
                "series_name": "DK SERIES",
                "series_subtitle": "End-Suction Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"DK100","powerHP":"1.5","powerKW":"1.1","maxHead":"22","maxFlow":"250 L/min","suctionHead":"5","inletOutlet":"2\" x 2\"","weight":"22.0","image":""},
                    {"model":"DK200","powerHP":"3.0","powerKW":"2.2","maxHead":"25","maxFlow":"400 L/min","suctionHead":"5","inletOutlet":"3\" x 3\"","weight":"38.0","image":""},
                ],
            },
            {
                "series_key": "shfm-series",
                "series_name": "SHFM SERIES",
                "series_subtitle": "Stainless Steel Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"SHFM150","powerHP":"1.5","powerKW":"1.1","maxHead":"32","maxFlow":"160 L/min","suctionHead":"7","inletOutlet":"1.5\" x 1.5\"","weight":"12.0","image":""},
                    {"model":"SHFM220","powerHP":"2.0","powerKW":"1.5","maxHead":"38","maxFlow":"220 L/min","suctionHead":"7","inletOutlet":"2\" x 2\"","weight":"16.5","image":""},
                ],
            },
            {
                "series_key": "scm-series",
                "series_name": "SCM SERIES",
                "series_subtitle": "Side-Channel Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"SCM150","powerHP":"1.5","powerKW":"1.1","maxHead":"35","maxFlow":"100 L/min","suctionHead":"6","inletOutlet":"1.5\" x 1.5\"","weight":"13.5","image":""},
                ],
            },
            {
                "series_key": "cm-series",
                "series_name": "CM SERIES",
                "series_subtitle": "Compact Monoblock Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"CM50","powerHP":"0.5","powerKW":"0.37","maxHead":"28","maxFlow":"50 L/min","suctionHead":"6","inletOutlet":"1\" x 1\"","weight":"6.0","image":""},
                    {"model":"CM75","powerHP":"0.75","powerKW":"0.55","maxHead":"35","maxFlow":"75 L/min","suctionHead":"6","inletOutlet":"1\" x 1\"","weight":"7.5","image":""},
                ],
            },
            {
                "series_key": "cpm-open-series",
                "series_name": "CPm OPEN SERIES",
                "series_subtitle": "Open Impeller Centrifugal Pumps",
                "image_hero": "/images/products/centrifugal-pump.png",
                "image_overview": "/images/products/centrifugal-pump.png",
                "models": [
                    {"model":"CPmO100","powerHP":"1.0","powerKW":"0.75","maxHead":"18","maxFlow":"200 L/min","suctionHead":"5","inletOutlet":"2\" x 2\"","weight":"14.0","image":""},
                    {"model":"CPmO150","powerHP":"1.5","powerKW":"1.1","maxHead":"22","maxFlow":"300 L/min","suctionHead":"5","inletOutlet":"2\" x 2\"","weight":"18.0","image":""},
                ],
            },
        ],
    },
    {
        "slug": "self-priming-jet-pump",
        "name": "Self-Priming Jet Pump（自吸喷射泵）",
        "series": [
            {
                "series_key": "gwp-series",
                "series_name": "GWPm SERIES",
                "series_subtitle": "Standard Self-Priming Jet Pumps",
                "image_hero": "/images/products/self-priming-jet-pump.png",
                "image_overview": "/images/products/self-priming-jet-pump.png",
                "models": [
                    {"model":"GWPm250","powerHP":"0.33","powerKW":"0.25","maxHead":"22","maxFlow":"25 L/min","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"4.2","image":""},
                    {"model":"GWPm370","powerHP":"0.5","powerKW":"0.37","maxHead":"28","maxFlow":"30 L/min","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"4.8","image":""},
                    {"model":"GWPm550","powerHP":"0.75","powerKW":"0.55","maxHead":"35","maxFlow":"45 L/min","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"6.5","image":""},
                    {"model":"GWPm750","powerHP":"1.0","powerKW":"0.75","maxHead":"42","maxFlow":"55 L/min","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"8.2","image":""},
                ],
            },
            {
                "series_key": "ajm-series",
                "series_name": "AJm SERIES",
                "series_subtitle": "Heavy-Duty Cast Iron Jet Pumps",
                "image_hero": "/images/products/self-priming-jet-pump.png",
                "image_overview": "/images/products/self-priming-jet-pump.png",
                "models": [
                    {"model":"AJm37","powerHP":"0.5","powerKW":"0.37","maxHead":"32","maxFlow":"40 L/min","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"8.5","image":""},
                    {"model":"AJm75","powerHP":"1.0","powerKW":"0.75","maxHead":"45","maxFlow":"70 L/min","suctionHead":"9","inletOutlet":"1\" x 1\"","weight":"12.0","image":""},
                    {"model":"AJm110","powerHP":"1.5","powerKW":"1.1","maxHead":"50","maxFlow":"90 L/min","suctionHead":"9","inletOutlet":"1.25\" x 1\"","weight":"15.5","image":""},
                    {"model":"AJm150","powerHP":"2.0","powerKW":"1.5","maxHead":"50","maxFlow":"100 L/min","suctionHead":"9","inletOutlet":"1.5\" x 1.25\"","weight":"19.0","image":""},
                ],
            },
            {
                "series_key": "jet-series",
                "series_name": "JET SERIES",
                "series_subtitle": "Classic Jet Pumps",
                "image_hero": "/images/products/self-priming-jet-pump.png",
                "image_overview": "/images/products/self-priming-jet-pump.png",
                "models": [
                    {"model":"JET100","powerHP":"1.0","powerKW":"0.75","maxHead":"40","maxFlow":"60 L/min","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"10.0","image":""},
                    {"model":"JET150","powerHP":"1.5","powerKW":"1.1","maxHead":"48","maxFlow":"80 L/min","suctionHead":"8","inletOutlet":"1.25\" x 1\"","weight":"13.5","image":""},
                ],
            },
            {
                "series_key": "jsw-series",
                "series_name": "JSW SERIES",
                "series_subtitle": "Shallow Well Jet Pumps",
                "image_hero": "/images/products/self-priming-jet-pump.png",
                "image_overview": "/images/products/self-priming-jet-pump.png",
                "models": [
                    {"model":"JSW100","powerHP":"1.0","powerKW":"0.75","maxHead":"45","maxFlow":"70 L/min","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"9.5","image":""},
                    {"model":"JSW150","powerHP":"1.5","powerKW":"1.1","maxHead":"55","maxFlow":"90 L/min","suctionHead":"8","inletOutlet":"1.25\" x 1\"","weight":"12.5","image":""},
                ],
            },
        ],
    },
    {
        "slug": "submersible-sewage-pump",
        "name": "Submersible Sewage Pump（潜水污水泵）",
        "series": [
            {
                "series_key": "wqm-series",
                "series_name": "WQm SERIES",
                "series_subtitle": "Submersible Sewage Pumps",
                "image_hero": "/images/products/submersible-sewage-pump.png",
                "image_overview": "/images/products/submersible-sewage-pump.png",
                "models": [
                    {"model":"WQm7-15-0.75","powerHP":"1.0","powerKW":"0.75","maxHead":"15","maxFlow":"7 m³/h","suctionHead":"—","inletOutlet":"2\" (50mm)","weight":"14.0","image":"","maxParticle":"35mm"},
                    {"model":"WQm7-22-1.1","powerHP":"1.5","powerKW":"1.1","maxHead":"22","maxFlow":"7 m³/h","suctionHead":"—","inletOutlet":"2\" (50mm)","weight":"18.0","image":"","maxParticle":"35mm"},
                    {"model":"WQm15-15-1.5","powerHP":"2.0","powerKW":"1.5","maxHead":"15","maxFlow":"15 m³/h","suctionHead":"—","inletOutlet":"3\" (80mm)","weight":"24.0","image":"","maxParticle":"50mm"},
                    {"model":"WQm25-10-2.2","powerHP":"3.0","powerKW":"2.2","maxHead":"10","maxFlow":"25 m³/h","suctionHead":"—","inletOutlet":"3\" (80mm)","weight":"32.0","image":"","maxParticle":"50mm"},
                ],
            },
            {
                "series_key": "wqkm-series",
                "series_name": "WQKm SERIES",
                "series_subtitle": "Heavy-Duty Submersible Sewage Pumps",
                "image_hero": "/images/products/submersible-sewage-pump.png",
                "image_overview": "/images/products/submersible-sewage-pump.png",
                "models": [
                    {"model":"WQKm10-22-2.2","powerHP":"3.0","powerKW":"2.2","maxHead":"22","maxFlow":"10 m³/h","suctionHead":"—","inletOutlet":"3\" (80mm)","weight":"38.0","image":"","maxParticle":"65mm"},
                    {"model":"WQKm20-22-3.0","powerHP":"4.0","powerKW":"3.0","maxHead":"22","maxFlow":"20 m³/h","suctionHead":"—","inletOutlet":"4\" (100mm)","weight":"52.0","image":"","maxParticle":"65mm"},
                ],
            },
            {
                "series_key": "wq-series",
                "series_name": "WQ SERIES",
                "series_subtitle": "Standard Sewage Pumps",
                "image_hero": "/images/products/submersible-sewage-pump.png",
                "image_overview": "/images/products/submersible-sewage-pump.png",
                "models": [
                    {"model":"WQ7-15-0.75","powerHP":"1.0","powerKW":"0.75","maxHead":"15","maxFlow":"7 m³/h","suctionHead":"—","inletOutlet":"2\" (50mm)","weight":"12.0","image":"","maxParticle":"30mm"},
                    {"model":"WQ20-15-1.5","powerHP":"2.0","powerKW":"1.5","maxHead":"15","maxFlow":"20 m³/h","suctionHead":"—","inletOutlet":"3\" (80mm)","weight":"22.0","image":"","maxParticle":"45mm"},
                ],
            },
            {
                "series_key": "qdx-series",
                "series_name": "QDX SERIES",
                "series_subtitle": "Compact Submersible Drainage Pumps",
                "image_hero": "/images/products/submersible-sewage-pump.png",
                "image_overview": "/images/products/submersible-sewage-pump.png",
                "models": [
                    {"model":"QDX10-16-0.75","powerHP":"1.0","powerKW":"0.75","maxHead":"16","maxFlow":"10 m³/h","suctionHead":"—","inletOutlet":"1.5\" (40mm)","weight":"8.5","image":"","maxParticle":"5mm"},
                    {"model":"QDX15-10-0.55","powerHP":"0.75","powerKW":"0.55","maxHead":"10","maxFlow":"15 m³/h","suctionHead":"—","inletOutlet":"1.5\" (40mm)","weight":"7.0","image":"","maxParticle":"5mm"},
                ],
            },
        ],
    },
    {
        "slug": "variable-frequency-pump",
        "name": "Variable Frequency Pump（永磁变频泵）",
        "series": [
            {
                "series_key": "vfm-series",
                "series_name": "VFm SERIES",
                "series_subtitle": "Variable Frequency Monoblock Pumps",
                "image_hero": "/images/products/variable-frequency-pump.png",
                "image_overview": "/images/products/variable-frequency-pump.png",
                "models": [
                    {"model":"VFm370","powerHP":"0.5","powerKW":"0.37","maxHead":"35","maxFlow":"2.5 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"9.5","image":""},
                    {"model":"VFm550","powerHP":"0.75","powerKW":"0.55","maxHead":"42","maxFlow":"3.5 m³/h","suctionHead":"7","inletOutlet":"1\" x 1\"","weight":"11.8","image":""},
                    {"model":"VFm750","powerHP":"1.0","powerKW":"0.75","maxHead":"48","maxFlow":"4.8 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"14.5","image":""},
                    {"model":"VFm1100","powerHP":"1.5","powerKW":"1.1","maxHead":"52","maxFlow":"6.0 m³/h","suctionHead":"8","inletOutlet":"1.25\" x 1\"","weight":"18.0","image":""},
                ],
            },
            {
                "series_key": "vfc-series",
                "series_name": "VFc SERIES",
                "series_subtitle": "Variable Frequency Centrifugal Pumps",
                "image_hero": "/images/products/variable-frequency-pump.png",
                "image_overview": "/images/products/variable-frequency-pump.png",
                "models": [
                    {"model":"VFc1100","powerHP":"1.5","powerKW":"1.1","maxHead":"38","maxFlow":"9.6 m³/h","suctionHead":"7","inletOutlet":"1.5\" x 1.5\"","weight":"18.5","image":""},
                    {"model":"VFc1500","powerHP":"2.0","powerKW":"1.5","maxHead":"44","maxFlow":"13.2 m³/h","suctionHead":"7","inletOutlet":"1.5\" x 1.5\"","weight":"23.0","image":""},
                    {"model":"VFc2200","powerHP":"3.0","powerKW":"2.2","maxHead":"50","maxFlow":"18.0 m³/h","suctionHead":"7","inletOutlet":"2\" x 2\"","weight":"30.0","image":""},
                ],
            },
            {
                "series_key": "t-series",
                "series_name": "T SERIES",
                "series_subtitle": "Twin Impeller Variable Frequency Pumps",
                "image_hero": "/images/products/variable-frequency-pump.png",
                "image_overview": "/images/products/variable-frequency-pump.png",
                "models": [
                    {"model":"T550","powerHP":"0.75","powerKW":"0.55","maxHead":"55","maxFlow":"3.0 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"13.5","image":""},
                    {"model":"T750","powerHP":"1.0","powerKW":"0.75","maxHead":"65","maxFlow":"4.2 m³/h","suctionHead":"8","inletOutlet":"1\" x 1\"","weight":"16.0","image":""},
                    {"model":"T1100","powerHP":"1.5","powerKW":"1.1","maxHead":"72","maxFlow":"5.5 m³/h","suctionHead":"8","inletOutlet":"1.25\" x 1\"","weight":"20.0","image":""},
                ],
            },
        ],
    },
    {
        "slug": "multi-stage-pump-sets",
        "name": "Multi-Stage Pump Sets（多级泵/泵组）",
        "series": [
            {
                "series_key": "cri-series",
                "series_name": "CRI SERIES",
                "series_subtitle": "Inline Multi-Stage Centrifugal Pumps",
                "image_hero": "/images/products/multi-stage-pump-sets.png",
                "image_overview": "/images/products/multi-stage-pump-sets.png",
                "models": [
                    {"model":"CRI1-8","powerHP":"0.75","powerKW":"0.55","maxHead":"52","maxFlow":"1.2 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"8.5","image":""},
                    {"model":"CRI3-10","powerHP":"1.5","powerKW":"1.1","maxHead":"72","maxFlow":"3.6 m³/h","suctionHead":"—","inletOutlet":"1.5\" x 1.5\"","weight":"14.0","image":""},
                    {"model":"CRI5-8","powerHP":"2.0","powerKW":"1.5","maxHead":"62","maxFlow":"6.0 m³/h","suctionHead":"—","inletOutlet":"1.5\" x 1.5\"","weight":"19.0","image":""},
                    {"model":"CRI10-5","powerHP":"3.0","powerKW":"2.2","maxHead":"45","maxFlow":"12.0 m³/h","suctionHead":"—","inletOutlet":"2\" x 2\"","weight":"28.0","image":""},
                ],
            },
            {
                "series_key": "bpsm-series",
                "series_name": "BPSM SERIES",
                "series_subtitle": "Booster Pump Sets with Motor",
                "image_hero": "/images/products/multi-stage-pump-sets.png",
                "image_overview": "/images/products/multi-stage-pump-sets.png",
                "models": [
                    {"model":"BPSM2-7","powerHP":"0.75","powerKW":"0.55","maxHead":"49","maxFlow":"2.4 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"9.0","image":""},
                    {"model":"BPSM2-12","powerHP":"1.5","powerKW":"1.1","maxHead":"85","maxFlow":"2.4 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"13.5","image":""},
                    {"model":"BPSM4-10","powerHP":"2.0","powerKW":"1.5","maxHead":"72","maxFlow":"4.8 m³/h","suctionHead":"—","inletOutlet":"1.5\" x 1.5\"","weight":"18.0","image":""},
                ],
            },
            {
                "series_key": "hmc-series",
                "series_name": "HMC SERIES",
                "series_subtitle": "Horizontal Multi-Stage Centrifugal Pumps",
                "image_hero": "/images/products/multi-stage-pump-sets.png",
                "image_overview": "/images/products/multi-stage-pump-sets.png",
                "models": [
                    {"model":"HMC-2-37","powerHP":"0.5","powerKW":"0.37","maxHead":"36","maxFlow":"2.0 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"11.0","image":""},
                    {"model":"HMC-2-55","powerHP":"0.75","powerKW":"0.55","maxHead":"54","maxFlow":"2.0 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"14.0","image":""},
                    {"model":"HMC-2-110","powerHP":"1.5","powerKW":"1.1","maxHead":"108","maxFlow":"2.0 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"21.5","image":""},
                    {"model":"HMC-4-75","powerHP":"1.0","powerKW":"0.75","maxHead":"72","maxFlow":"4.0 m³/h","suctionHead":"—","inletOutlet":"1.25\" x 1\"","weight":"18.0","image":""},
                    {"model":"HMC-8-75","powerHP":"1.0","powerKW":"0.75","maxHead":"54","maxFlow":"8.0 m³/h","suctionHead":"—","inletOutlet":"1.5\" x 1.5\"","weight":"22.0","image":""},
                    {"model":"HMC-15-110","powerHP":"1.5","powerKW":"1.1","maxHead":"54","maxFlow":"15.0 m³/h","suctionHead":"—","inletOutlet":"2\" x 2\"","weight":"32.0","image":""},
                ],
            },
            {
                "series_key": "cdlf-series",
                "series_name": "CDLF SERIES",
                "series_subtitle": "Vertical Inline Multi-Stage Pumps",
                "image_hero": "/images/products/multi-stage-pump-sets.png",
                "image_overview": "/images/products/multi-stage-pump-sets.png",
                "models": [
                    {"model":"CDLF2-5","powerHP":"0.5","powerKW":"0.37","maxHead":"38","maxFlow":"2.0 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"10.5","image":""},
                    {"model":"CDLF2-10","powerHP":"0.75","powerKW":"0.55","maxHead":"68","maxFlow":"2.0 m³/h","suctionHead":"—","inletOutlet":"1\" x 1\"","weight":"14.0","image":""},
                    {"model":"CDLF4-8","powerHP":"1.5","powerKW":"1.1","maxHead":"72","maxFlow":"4.0 m³/h","suctionHead":"—","inletOutlet":"1.5\" x 1.5\"","weight":"22.0","image":""},
                ],
            },
        ],
    },
]


SHEET_NAMES = {
    "peripheral-pump":        "漩涡泵 Peripheral",
    "centrifugal-pump":       "离心泵 Centrifugal",
    "self-priming-jet-pump":  "喷射泵 JetPump",
    "submersible-sewage-pump":"污水泵 Sewage",
    "variable-frequency-pump":"变频泵 VF",
    "multi-stage-pump-sets":  "多级泵 MultiStage",
}

def build_sheet(wb: Workbook, cat: dict):
    title = SHEET_NAMES.get(cat["slug"], cat["slug"][:31])
    ws = wb.create_sheet(title=title)
    ws.freeze_panes = "A4"

    # ── Legend row ──────────────────────────────────────────────────
    leg_row = 1
    ws.row_dimensions[leg_row].height = 22
    lc = ws.cell(row=leg_row, column=1,
        value="🟡 黄色单元格 = 需要填写/修改的数据   🟢 绿色单元格 = 图片文件名（如 pump-photo.jpg）   白色 = 系列描述文字")
    lc.font = Font(name="Arial", size=9, color="333333")
    lc.fill = PatternFill("solid", fgColor="FFFDE7")
    lc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=20)

    # ── Title row ────────────────────────────────────────────────────
    title_row = 2
    ws.row_dimensions[title_row].height = 28
    tc = ws.cell(row=title_row, column=1,
                 value=f"ALFAGRAND — {cat['name']} — 产品数据填写表")
    tc.font = Font(name="Arial", bold=True, size=14, color=C_TITLE_FG)
    tc.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=20)

    # ── Column header row ────────────────────────────────────────────
    col_row = 3
    ws.row_dimensions[col_row].height = 40

    COLS = [
        ("系列",          "Series Key\n(勿修改)", 16),
        ("系列名称",       "Series Name\n(e.g. PM SERIES)", 18),
        ("系列副标题",     "Subtitle\n(简短描述)", 22),
        ("Hero图片",      "Hero Image\n文件名", 22),
        ("概览图片",       "Overview Image\n文件名", 22),
        ("型号",          "Model", 14),
        ("功率HP",        "Power\n(HP)", 10),
        ("功率kW",        "Power\n(kW)", 10),
        ("最大扬程m",      "Max Head\n(m)", 12),
        ("最大流量",       "Max Flow\n(m³/h or L/min)", 18),
        ("吸程m",         "Suction Head\n(m，无填—)", 14),
        ("进出口",        "Inlet/Outlet\n(inch or mm)", 18),
        ("重量kg",        "G.W.\n(kg)", 10),
        ("型号图片",      "Model Photo\n文件名", 22),
        ("最大颗粒mm",    "Max Particle\n(mm，污水泵填)", 16),
        ("备注",          "备注 / Remark", 20),
    ]

    for i, (cn, en, w) in enumerate(COLS, start=1):
        header_cell(ws, col_row, i, f"{cn}\n{en}", wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ────────────────────────────────────────────────────
    cur_row = col_row + 1
    for series in cat["series"]:
        first_model = True
        for m in series["models"]:
            ws.row_dimensions[cur_row].height = 18
            row_bg = C_ROW_A if (cur_row % 2 == 0) else C_ROW_B

            # A: series key (read-only, dark)
            data_cell(ws, cur_row, 1, series["series_key"], bg="0A1829", fg="5599CC")

            # B-C: series name / subtitle (yellow: fill if wrong)
            input_cell(ws, cur_row, 2, series["series_name"] if first_model else "")
            input_cell(ws, cur_row, 3, series["series_subtitle"] if first_model else "")

            # D-E: images (green)
            image_cell(ws, cur_row, 4, series["image_hero"] if first_model else "")
            image_cell(ws, cur_row, 5, series["image_overview"] if first_model else "")

            # F: model
            input_cell(ws, cur_row, 6, m["model"])

            # G-M: numeric specs
            for col_i, key in enumerate(["powerHP","powerKW","maxHead","maxFlow","suctionHead","inletOutlet","weight"], start=7):
                input_cell(ws, cur_row, col_i, m.get(key, ""))

            # N: model photo (green)
            image_cell(ws, cur_row, 14, m.get("image", ""))

            # O: max particle
            input_cell(ws, cur_row, 15, m.get("maxParticle", ""))

            # P: remark
            input_cell(ws, cur_row, 16, "")

            first_model = False
            cur_row += 1

        # blank separator row between series
        ws.row_dimensions[cur_row].height = 8
        for col_i in range(1, 17):
            c = ws.cell(row=cur_row, column=col_i, value="")
            c.fill = PatternFill("solid", fgColor="050D18")
        cur_row += 1

    return ws


def build_guide_sheet(wb: Workbook):
    ws = wb.create_sheet(title="📋 使用说明", index=0)
    ws.sheet_view.showGridLines = False

    rows = [
        ("ALFAGRAND 产品数据填写模板 — 使用说明", C_TITLE_BG, C_TITLE_FG, 16, True),
        ("", C_TITLE_BG, C_TITLE_FG, 6, False),
        ("📌 总体流程", "0D3047", "00E5FF", 12, True),
        ("1. 此文件共 7 个 Sheet：1个说明页 + 6个品类数据页", C_ROW_B, "D0E8FF", 10, False),
        ("2. 在各品类 Sheet 中，按系列逐行修改型号参数", C_ROW_B, "D0E8FF", 10, False),
        ("3. 可以增加/删除型号行（在同一系列块内）", C_ROW_B, "D0E8FF", 10, False),
        ("4. 图片文件单独收集好，统一发给AI填充", C_ROW_B, "D0E8FF", 10, False),
        ("5. 填好后把此Excel发回，AI会自动更新网站代码", C_ROW_B, "D0E8FF", 10, False),
        ("", C_ROW_B, "D0E8FF", 6, False),
        ("🟡 黄色单元格说明", "0D3047", "FFD700", 12, True),
        ("所有黄色背景单元格都需要检查/修改", C_ROW_B, "D0E8FF", 10, False),
        ("包括：系列名称、副标题、全部型号参数", C_ROW_B, "D0E8FF", 10, False),
        ("", C_ROW_B, "D0E8FF", 6, False),
        ("🟢 绿色单元格说明（图片）", "0D3047", "88FF88", 12, True),
        ("填写图片文件名，格式示例：peripheral-pump-pm45.jpg", C_ROW_B, "D0E8FF", 10, False),
        ("图片统一收集后发给AI，AI会自动上传到正确目录", C_ROW_B, "D0E8FF", 10, False),
        ("支持格式：.jpg / .jpeg / .png / .webp", C_ROW_B, "D0E8FF", 10, False),
        ("", C_ROW_B, "D0E8FF", 6, False),
        ("📊 参数字段说明", "0D3047", "00E5FF", 12, True),
        ("Series Key     - 系列路由键（勿修改，影响URL）", C_ROW_B, "D0E8FF", 10, False),
        ("Series Name    - 系列展示名称，如 'PM SERIES'", C_ROW_B, "D0E8FF", 10, False),
        ("Subtitle       - 系列副标题，一句话描述", C_ROW_B, "D0E8FF", 10, False),
        ("Hero Image     - 系列页顶部展示图（同系列可共用一张）", C_ROW_B, "D0E8FF", 10, False),
        ("Overview Image - 系列概览区图片（可与Hero图相同）", C_ROW_B, "D0E8FF", 10, False),
        ("Model          - 型号名，如 'PM45A'（保持英文大写）", C_ROW_B, "D0E8FF", 10, False),
        ("Power HP/kW    - 功率（如 0.37 kW = 0.5 HP）", C_ROW_B, "D0E8FF", 10, False),
        ("Max Head       - 最大扬程（米）", C_ROW_B, "D0E8FF", 10, False),
        ("Max Flow       - 最大流量（注意单位：m³/h 或 L/min）", C_ROW_B, "D0E8FF", 10, False),
        ("Suction Head   - 吸程（米），无吸程填 —", C_ROW_B, "D0E8FF", 10, False),
        ("Inlet/Outlet   - 进出口径，如 '1\" x 1\"' 或 '50mm'", C_ROW_B, "D0E8FF", 10, False),
        ("G.W.           - 毛重（千克）", C_ROW_B, "D0E8FF", 10, False),
        ("Model Photo    - 该型号专属图片（可为空，空=用系列图）", C_ROW_B, "D0E8FF", 10, False),
        ("Max Particle   - 最大过颗粒（mm），仅污水泵需填", C_ROW_B, "D0E8FF", 10, False),
        ("", C_ROW_B, "D0E8FF", 6, False),
        ("⚠️ 注意事项", "0D3047", "FF8C00", 12, True),
        ("- 不要修改深色的 'Series Key' 列，否则会导致URL 404", C_ROW_B, "FFCCAA", 10, False),
        ("- 如需新增系列，请告知AI，由AI手动添加路由", C_ROW_B, "FFCCAA", 10, False),
        ("- 流量单位请保持一致（同一品类全用 m³/h 或全用 L/min）", C_ROW_B, "FFCCAA", 10, False),
        ("- 型号名请保持英文+数字，不要含特殊字符", C_ROW_B, "FFCCAA", 10, False),
    ]

    ws.column_dimensions["A"].width = 80

    for i, (text, bg, fg, size, bold) in enumerate(rows, start=1):
        ws.row_dimensions[i].height = 20 if size >= 12 else 16
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", size=size, color=fg, bold=bold)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=0 if bold else 1)


def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_guide_sheet(wb)
    for cat in CATEGORIES:
        build_sheet(wb, cat)

    wb.save(OUTPUT)
    print("Saved: " + OUTPUT)

if __name__ == "__main__":
    main()
