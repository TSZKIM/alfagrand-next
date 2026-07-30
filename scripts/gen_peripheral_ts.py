import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'D:/HuaweiMoveData/Users/华为/Desktop/alfagrand/ALFAGRAND_Product_Data_Completed.xlsx', data_only=True)
sheetnames = wb.sheetnames
ws = wb[sheetnames[1]]
rows = list(ws.iter_rows(values_only=True))

pic_map = {
    'pm-series':     'PM.png',
    'pkm-series':    'pkm.png',
    'ps-series':     'PS-GP.png',
    'gp-series':     'PS-GP.png',
    'qb-series':     'QB.png',
    '1wzb-series':   '1WZB.png',
    'awzb-series':   'AWZB.png',
    'pw-series':     'PW.png',
    'autoqb-series': 'AUTOQB.png',
}

def safe(v):
    if v is None: return ''
    s = str(v).strip()
    if s in ('None', 'nan', '') or s.startswith('(') or s.startswith('/images/products/peripheral'):
        return ''
    return s

series_data = {}
series_order = []
cur_key = None
cur_name = ''
cur_subtitle = ''

for row in rows[3:]:
    k = safe(row[0])
    k_lower = k.lower().replace(' ', '-') if k else ''
    if k:
        norm_key = k_lower
        if norm_key not in series_data:
            sname = safe(row[1]) or cur_name
            ssub = safe(row[2]) or cur_subtitle
            series_data[norm_key] = {'name': sname, 'subtitle': ssub, 'models': []}
            series_order.append(norm_key)
        cur_key = norm_key
        if safe(row[1]):
            cur_name = safe(row[1])
            series_data[cur_key]['name'] = cur_name
        if safe(row[2]):
            cur_subtitle = safe(row[2])
            series_data[cur_key]['subtitle'] = cur_subtitle

    model = safe(row[5])
    if model and cur_key:
        series_data[cur_key]['models'].append({
            'model': model,
            'powerHP': safe(row[6]),
            'powerKW': safe(row[7]),
            'maxHead': safe(row[8]),
            'maxFlow': safe(row[9]),
            'suctionHead': safe(row[10]),
            'inletOutlet': safe(row[11]),
            'weight': safe(row[12]),
        })

def esc(s):
    return str(s).replace('\\', '\\\\').replace('`', '\\`')

def gen_block(key, data):
    name = data['name']
    subtitle = data['subtitle']
    models = data['models']
    pic = pic_map.get(key, 'PM.png')
    img = f'/images/products/{pic}'
    n_models = len(models)

    models_ts = ''
    for m in models:
        models_ts += '      {\n'
        models_ts += f'        model: `{esc(m["model"])}`,\n'
        models_ts += f'        powerHP: `{esc(m["powerHP"])}`,\n'
        models_ts += f'        powerKW: `{esc(m["powerKW"])}`,\n'
        models_ts += f'        maxHead: `{esc(m["maxHead"])}`,\n'
        models_ts += f'        maxFlow: `{esc(m["maxFlow"])}`,\n'
        models_ts += f'        suctionHead: `{esc(m["suctionHead"])}`,\n'
        models_ts += f'        inletOutlet: `{esc(m["inletOutlet"])}`,\n'
        models_ts += f'        weight: `{esc(m["weight"])}`,\n'
        models_ts += '      },\n'

    head_min = models[0]["maxHead"]
    head_max = models[-1]["maxHead"]

    b = f'  /* \u2500\u2500\u2500 {name} \u2500\u2500\u2500 */\n'
    b += f'  "peripheral-pump/{key}": {{\n'
    b += f'    title: `{esc(name)}`,\n'
    b += f'    subtitle: `{esc(subtitle)}`,\n'
    b += f'    desc: `The {esc(name)} delivers reliable performance for domestic water supply and light commercial applications. Compact design with strong suction capability and easy installation.`,\n'
    b += '    heroCheckmarks: [\n'
    b += '      `Cast iron body with stainless steel shaft`,\n'
    b += '      `Self-priming up to 8m`,\n'
    b += '      `Suitable for clean water`,\n'
    b += '      `CE certified`,\n'
    b += '    ],\n'
    b += '    modelImages: [\n'
    b += f'      {{ src: `{img}`, label: `{esc(name)}` }},\n'
    b += '    ],\n'
    b += '    stats: [\n'
    b += f'      {{ icon: "\U0001f4e6", label: "Models", value: "{n_models}+" }},\n'
    b += '      { icon: "\U0001f3ed", label: "CE / ISO", value: "Certified" },\n'
    b += '      { icon: "\u26a1", label: "50/60 Hz", value: "Available" },\n'
    b += '    ],\n'
    b += f'    overviewTitle: `{esc(name)} Overview`,\n'
    b += f'    overviewDesc: `The {esc(name)} ({esc(subtitle)}) is designed for reliable domestic water supply. Features vortex/peripheral pump technology for stable pressure.`,\n'
    b += f'    overviewImage: `{img}`,\n'
    b += '    phaseOptions: [\n'
    b += '      { label: "Single Phase", desc: "1~ 220V 50Hz/60Hz" },\n'
    b += '      { label: "Three Phase", desc: "3~ 380V 50Hz/60Hz" },\n'
    b += '    ],\n'
    b += '    tableCols: [\n'
    b += '      { key: "model",       label: "Model" },\n'
    b += '      { key: "powerHP",     label: "Power (HP)" },\n'
    b += '      { key: "powerKW",     label: "Power (kW)" },\n'
    b += '      { key: "maxHead",     label: "Max Head (m)" },\n'
    b += '      { key: "maxFlow",     label: "Max Flow" },\n'
    b += '      { key: "suctionHead", label: "Suction Head (m)" },\n'
    b += '      { key: "inletOutlet", label: "Inlet/Outlet" },\n'
    b += '      { key: "weight",      label: "Weight (kg)" },\n'
    b += '    ],\n'
    b += '    models: [\n'
    b += models_ts
    b += '    ],\n'
    b += '    performanceCurves: [],\n'
    b += '    features: [\n'
    b += '      { icon: "\u2699\ufe0f", title: "Peripheral Pump Design", desc: "Vortex impeller delivers stable pressure across the full flow range." },\n'
    b += '      { icon: "\U0001f512", title: "Thermal Protection", desc: "Built-in thermal overload protects the motor from damage." },\n'
    b += '      { icon: "\U0001f3c6", title: "CE Certified", desc: "Meets European safety and performance standards." },\n'
    b += '      { icon: "\U0001f527", title: "Easy Maintenance", desc: "Simple structure allows quick disassembly and servicing." },\n'
    b += '    ],\n'
    b += '    applications: [\n'
    b += '      { icon: "\U0001f3e0", label: "Domestic Water Supply", desc: "Ideal for home pressure boosting and water transfer." },\n'
    b += '      { icon: "\U0001f33f", label: "Garden Irrigation", desc: "Supplies garden sprinkler and drip irrigation systems." },\n'
    b += '      { icon: "\U0001f3d7\ufe0f", label: "Light Industrial", desc: "Suitable for small factories and workshops." },\n'
    b += '      { icon: "\U0001f4a7", label: "Pressure Boosting", desc: "Enhances water pressure in multi-story buildings." },\n'
    b += '    ],\n'
    b += '    faqs: [\n'
    b += f'      {{ q: "What is the max head of {esc(name)}?", a: "Depending on the model, max head ranges from {head_min}m to {head_max}m." }},\n'
    b += '      { q: "Is it self-priming?", a: "Yes, all peripheral pumps in this series are self-priming up to 8m." },\n'
    b += '    ],\n'
    b += f'    ctaTitle: `Ready to Order {esc(name)}?`,\n'
    b += f'    ctaDesc: `Contact our team for pricing, technical specs, and OEM/ODM options.`,\n'
    b += '  },\n\n'
    return b

new_peripheral = ''
for key in series_order:
    new_peripheral += gen_block(key, series_data[key])

ts_file = r'C:\Users\华为\.qclaw\workspace-37i6raipm851ul5j\alfagrand-next\src\data\series-data.ts'
content = open(ts_file, encoding='utf-8').read()

start_marker = '  /* \u2500\u2500\u2500 PM SERIES \u2500\u2500\u2500 */'
end_marker = '  /* \u2500\u2500\u2500 CPM SERIES \u2500\u2500\u2500 */'

start_idx = content.find(start_marker)
end_idx = content.find(end_marker)

if start_idx == -1 or end_idx == -1:
    print(f'ERROR: Markers not found. start={start_idx}, end={end_idx}')
    # debug
    pm_idx = content.find('"peripheral-pump/pm-series"')
    cpm_idx = content.find('"centrifugal-pump/cpm-series"')
    print(f'  pm-series at char {pm_idx}, cpm-series at char {cpm_idx}')
    # show context around pm-series
    if pm_idx >= 0:
        print('Context before pm-series:', repr(content[max(0,pm_idx-80):pm_idx]))
else:
    new_content = content[:start_idx] + new_peripheral + content[end_idx:]
    open(ts_file, 'w', encoding='utf-8').write(new_content)
    print(f'Done! {new_content.count(chr(10))} lines')
    print(f'Peripheral: {len(series_order)} series, {sum(len(v["models"]) for v in series_data.values())} models')
    for k in series_order:
        print(f'  {k}: {len(series_data[k]["models"])} models')
