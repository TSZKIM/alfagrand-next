#!/usr/bin/env python3
"""
parse_excel_to_ts.py  v2
读取 ALFAGRAND_Product_Data_Completed.xlsx，生成新的 src/data/series-data.ts
"""
import openpyxl
import os
import shutil
from pathlib import Path

EXCEL_PATH = r"D:\HuaweiMoveData\Users\华为\Desktop\alfagrand\ALFAGRAND_Product_Data_Completed.xlsx"
PIC_ROOT   = r"D:\HuaweiMoveData\Users\华为\Desktop\alfagrand\PIC"
OUT_TS     = r"C:\Users\华为\.qclaw\workspace-37i6raipm851ul5j\alfagrand-next\src\data\series-data.ts"
IMG_DST    = r"C:\Users\华为\.qclaw\workspace-37i6raipm851ul5j\alfagrand-next\public\images\products"

# ── Sheet → category slug ──────────────────────────────────────────────────
SHEET_CATEGORY = {
    "漩涡泵 Peripheral":       "peripheral-pump",
    "离心泵 Centrifugal":      "centrifugal-pump",
    "喷射泵 JetPump":          "self-priming-jet-pump",
    "污水泵 Sewage":            "submersible-sewage-pump",
    "变频泵 VFD Pump":         "variable-frequency-pump",
    "多级泵 MultiStage":       "multi-stage-pump-sets",
}

# ── PIC folder → category ───────────────────────────────────────────────────
PIC_FOLDER = {
    "peripheral-pump":          "Peripheral",
    "centrifugal-pump":         "Centrigual",
    "self-priming-jet-pump":    "Self-priming Jet",
    "submersible-sewage-pump":  "Submersible Sweage",
    "variable-frequency-pump":  "VFD",
    "multi-stage-pump-sets":    "Mutil-stage",
}

# ── Series key → best matching image filename (from PIC folder) ────────────
# Maps series_key → image filename (without extension, case-insensitive match)
SERIES_IMAGE_MAP = {
    # Peripheral
    "pm-series":          "PM",
    "Pkm-series":         "pkm",
    "PS-series":          "PS",
    "GP-series":          "PS-GP",
    "qb-series":          "QB",
    "1wzb-series":        "1WZB",
    "awzb-series":        "AWZB",
    "PW-series":          "PW",
    "autoqb series":      "AUTOQB",
    "2cpm-series":        "PM",
    "ae-series":          "PM",
    "as-s-series":        "PM",
    "gdwm-series":        "PM",
    "ghfm-series":        "PM",
    "gjsm-series":        "PM",
    "gpm-series":         "PM",
    "jswm-series":        "PM",
    "pq-series":          "PM",
    "pump-station-series":"PM",
    "zgd-series":         "PM",
    # Centrifugal
    "cpm-series":         "CPM",
    "dk-series":          "DK",
    "cp-series":          "CP",
    "hfm-series":         "HFM",
    "hct-series":         "HCT",
    "scm-series":         "SCM",
    "hnf-series":         "HNF",
    "2cp-series":         "2CP",
    # Jet Pump
    "jeta-series":        "JET-A",
    "jetb/c-series":      "JET-B",
    "jsp-series":         "JSP",
    "kjm-series":         "KJM",
    "jsw-series":         "JSW",
    "sjet-series":        "SJET",
    "dp-series":          "DP",
    "autojet-series":     "AUTOJET",
    # Sewage
    "qdx-series":         "QDX",
    "wqd-series":         "WQD",
    "V-series":           "QD",
    "qd-series":          "QD",
    "wq-series":          "WQ",
    # VFD
    "aigp-series":        "AIGP",
    "cmb-series":         "CMB",
    "t-series":           "T",
    # MultiStage
    "cht-series":         "CHT",
    "cdlf-series":        "CDLF",
    "chi-series":         "CHI",
    "mcs-series":         "MCS",
}

os.makedirs(IMG_DST, exist_ok=True)

def safe(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    bad = {"None", "nan", "-", "—", "N/A", "(e.g. my-pump.jpg)", ""}
    if s in bad or s.startswith("(e.g.") or s.startswith("/images/products/peripheral") \
       or s.startswith("/images/products/centrifugal") \
       or s.startswith("/images/products/self-priming") \
       or s.startswith("/images/products/submersible") \
       or s.startswith("/images/products/variable") \
       or s.startswith("/images/products/multi-stage"):
        return ""
    return s

def find_pic(stem: str, folder_name: str) -> str | None:
    """在 PIC/folder_name 中找到 stem 对应的图片（不区分大小写）"""
    folder_path = os.path.join(PIC_ROOT, folder_name)
    if not os.path.isdir(folder_path):
        return None
    for f in os.listdir(folder_path):
        name_no_ext = os.path.splitext(f)[0]
        if name_no_ext.lower() == stem.lower():
            return os.path.join(folder_path, f)
    return None

def get_series_image(series_key: str, cat_slug: str) -> str:
    """返回系列的 web 图片路径，同时确保文件已拷贝"""
    folder_name = PIC_FOLDER.get(cat_slug, "")
    stem = SERIES_IMAGE_MAP.get(series_key, "")
    if not stem:
        return "/images/products/placeholder.png"
    src = find_pic(stem, folder_name)
    if src:
        fname = os.path.basename(src)
        dst = os.path.join(IMG_DST, fname)
        if not os.path.exists(dst):
            shutil.copy2(src, dst)
        return f"/images/products/{fname}"
    return "/images/products/placeholder.png"

# ── Copy ALL images from PIC ────────────────────────────────────────────────
print("=== Copying all images ===")
for cat_slug, folder_name in PIC_FOLDER.items():
    folder_path = os.path.join(PIC_ROOT, folder_name)
    if os.path.isdir(folder_path):
        for f in os.listdir(folder_path):
            if f.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
                src_path = os.path.join(folder_path, f)
                dst_path = os.path.join(IMG_DST, f)
                if not os.path.exists(dst_path):
                    shutil.copy2(src_path, dst_path)
                    print(f"  COPY {f}")
                else:
                    print(f"  SKIP {f}")

# ── Parse workbook ──────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print("\nSheets:", wb.sheetnames)

# category_slug → ordered list of (series_key, series_info_dict)
all_data: dict[str, list] = {}

for sheet_name, cat_slug in SHEET_CATEGORY.items():
    if sheet_name not in wb.sheetnames:
        print(f"WARNING: Sheet '{sheet_name}' not found")
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # rows[0]=instruction, rows[1]=title, rows[2]=header, rows[3..]=data
    series_ordered: list[tuple[str, dict]] = []   # preserve insertion order
    series_lookup: dict[str, dict] = {}
    current_key = None

    for row in rows[3:]:
        if not any(v for v in row):
            continue  # skip empty rows

        s_key    = safe(row[0]) if len(row) > 0 else ""
        s_title  = safe(row[1]) if len(row) > 1 else ""
        s_sub    = safe(row[2]) if len(row) > 2 else ""
        # cols 3,4 = hero/overview images (often placeholder — we'll use series image map)
        model    = safe(row[5]) if len(row) > 5 else ""
        pHP      = safe(row[6]) if len(row) > 6 else ""
        pKW      = safe(row[7]) if len(row) > 7 else ""
        maxH     = safe(row[8]) if len(row) > 8 else ""
        maxF     = safe(row[9]) if len(row) > 9 else ""
        suct     = safe(row[10]) if len(row) > 10 else ""
        io_      = safe(row[11]) if len(row) > 11 else ""
        wt       = safe(row[12]) if len(row) > 12 else ""
        maxP     = safe(row[14]) if len(row) > 14 else ""

        if s_key:
            current_key = s_key
            if s_key not in series_lookup:
                d = {
                    "title":    s_title,
                    "subtitle": s_sub,
                    "models":   [],
                }
                series_lookup[s_key] = d
                series_ordered.append((s_key, d))
            else:
                # Update title/sub if now filled in
                if s_title:
                    series_lookup[s_key]["title"] = s_title
                if s_sub:
                    series_lookup[s_key]["subtitle"] = s_sub

        if current_key is None:
            continue

        if model:
            series_lookup[current_key]["models"].append({
                "model":       model,
                "powerHP":     pHP,
                "powerKW":     pKW,
                "maxHead":     maxH,
                "maxFlow":     maxF,
                "suctionHead": suct,
                "inletOutlet": io_,
                "weight":      wt,
                "maxParticle": maxP,
            })

    print(f"\n[{sheet_name}] → {cat_slug}: {len(series_ordered)} series")
    for sk, sv in series_ordered:
        print(f"  {sk}: {len(sv['models'])} models  title={sv['title']}")

    all_data[cat_slug] = series_ordered

# ── Defaults ────────────────────────────────────────────────────────────────
DEFAULT_FEATURES = {
    "peripheral-pump": [
        ("⚡", "Energy Efficient",  "Optimized hydraulic design minimizes energy consumption."),
        ("🔧", "Easy Maintenance",  "Simple structure with minimal moving parts for easy servicing."),
        ("💧", "Quiet Operation",   "Balanced impeller design ensures smooth, low-noise performance."),
        ("🏗️", "Compact Design",    "Space-saving form factor for residential and commercial use."),
    ],
    "centrifugal-pump": [
        ("⚡", "High Efficiency",   "Advanced hydraulic design for maximum energy efficiency."),
        ("💪", "Robust Build",      "Cast iron/stainless construction for long service life."),
        ("🌊", "High Flow Rate",    "Designed for continuous high-volume fluid transfer."),
        ("🔧", "Versatile",         "Wide range of applications from domestic to industrial."),
    ],
    "self-priming-jet-pump": [
        ("🚀", "Self-Priming",      "Automatically primes without manual water filling."),
        ("📐", "Deep Suction",      "Can draw water from wells up to 8–9m depth."),
        ("💧", "Clean Water",       "Designed for clean, lightly contaminated water handling."),
        ("🔧", "Easy Install",      "Simple piping connections for quick installation."),
    ],
    "submersible-sewage-pump": [
        ("♻️", "Anti-Clogging",     "Special impeller design handles solids and fibrous materials."),
        ("🛡️", "Submersible Sealed","Fully sealed motor for safe underwater operation."),
        ("💪", "High Durability",   "Cast iron construction resists corrosion and wear."),
        ("⚡", "Efficient Motor",   "High-efficiency motor reduces operating costs."),
    ],
    "variable-frequency-pump": [
        ("⚡", "Variable Speed",    "Inverter-controlled motor adapts to demand automatically."),
        ("💰", "Energy Saving",     "Up to 50% energy savings vs fixed-speed pumps."),
        ("🔇", "Ultra Quiet",       "Smooth speed variation eliminates water hammer noise."),
        ("📱", "Smart Control",     "Constant pressure control with built-in inverter."),
    ],
    "multi-stage-pump-sets": [
        ("📈", "High Pressure",     "Multi-stage impellers deliver high head pressure output."),
        ("🏗️", "Modular Design",    "Stages configured for specific pressure requirements."),
        ("🔧", "Low Maintenance",   "Reliable construction for continuous operation."),
        ("🌊", "Wide Application",  "HVAC, booster systems, and water treatment."),
    ],
}

DEFAULT_APPLICATIONS = {
    "peripheral-pump": [
        ("🏠", "Residential Boosting", "Household water pressure boosting"),
        ("🌱", "Garden Irrigation",    "Lawn and garden watering systems"),
        ("🏢", "Light Commercial",     "Small commercial water supply"),
        ("🏊", "Pool Circulation",     "Swimming pool and spa systems"),
    ],
    "centrifugal-pump": [
        ("🏭", "Industrial Transfer",  "Industrial fluid transfer applications"),
        ("🌾", "Agricultural",         "Irrigation and farm water supply"),
        ("🏢", "Commercial HVAC",      "Heating and cooling water circulation"),
        ("💧", "Water Treatment",      "Municipal water treatment plants"),
    ],
    "self-priming-jet-pump": [
        ("🏠", "Domestic Wells",       "Residential shallow well water supply"),
        ("🌱", "Garden & Farm",        "Irrigation from open water sources"),
        ("🏕️", "Remote Supply",        "Remote water supply for cabins"),
        ("🚿", "Pressure Boosting",    "Low-pressure water system boosting"),
    ],
    "submersible-sewage-pump": [
        ("🏘️", "Municipal Sewage",     "Urban sewage and wastewater transfer"),
        ("🏭", "Industrial Effluent",  "Factory wastewater handling"),
        ("🌧️", "Flood Control",        "Basement and flood drainage"),
        ("🏗️", "Construction",         "Construction site dewatering"),
    ],
    "variable-frequency-pump": [
        ("🏢", "Building Pressure",    "Constant pressure water supply in buildings"),
        ("🌾", "Smart Irrigation",     "Variable demand irrigation systems"),
        ("🏭", "Industrial Process",   "Process water with variable demand"),
        ("💧", "Water Treatment",      "Variable flow water treatment plants"),
    ],
    "multi-stage-pump-sets": [
        ("🏢", "High-Rise Boosting",   "Water pressure boosting for tall buildings"),
        ("🏭", "Industrial Process",   "High pressure industrial processes"),
        ("🔥", "Boiler Feed",          "Boiler feed and steam generation"),
        ("💧", "RO Systems",           "Reverse osmosis and filtration systems"),
    ],
}

DEFAULT_FAQS = [
    ("What is the warranty period?",
     "All ALFAGRAND pumps come with a 12-month warranty covering manufacturing defects."),
    ("Are spare parts available?",
     "Yes, we maintain a full inventory of spare parts for all models in our product range."),
    ("Do you offer OEM/ODM services?",
     "Yes, we offer OEM and ODM services for bulk orders. Contact us for details."),
]

def esc(s: str) -> str:
    return s.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")

# ── Generate TypeScript ──────────────────────────────────────────────────────
print("\n\n=== Generating series-data.ts ===")

L = []  # lines

header = '''/* ── Data Types ── */
export interface ModelVariant {
  model: string;
  powerHP: string;
  powerKW: string;
  maxHead: string;
  maxFlow: string;
  suctionHead?: string;
  inletOutlet: string;
  weight: string;
  maxParticle?: string;
}

export interface SeriesPerformancePoint {
  model: string;
  color: string;
  data: { flow: number; head: number }[];
}

export interface SeriesFeature {
  icon: string;
  title: string;
  desc: string;
}

export interface ApplicationCard {
  icon: string;
  label: string;
  desc: string;
}

export interface SeriesFAQ {
  q: string;
  a: string;
}

export interface SeriesData {
  title: string;
  subtitle: string;
  desc: string;
  heroCheckmarks: string[];
  modelImages: { src: string; label: string }[];
  stats: { icon: string; label: string; value: string }[];
  overviewTitle: string;
  overviewDesc: string;
  overviewImage: string;
  phaseOptions: { label: string; desc: string }[];
  tableCols: { key: string; label: string }[];
  models: ModelVariant[];
  performanceCurves: SeriesPerformancePoint[];
  features: SeriesFeature[];
  applications: ApplicationCard[];
  faqs: SeriesFAQ[];
  ctaTitle: string;
  ctaDesc: string;
}

/* ── Category name resolver ── */
export const categoryNames: Record<string, string> = {
  "peripheral-pump": "Peripheral Pump",
  "centrifugal-pump": "Centrifugal Pump",
  "self-priming-jet-pump": "Self-Priming Jet Pump",
  "submersible-sewage-pump": "Submersible Sewage Pump",
  "deep-well-pump": "Deep Well Pump",
  "variable-frequency-pump": "Variable Frequency Pump",
  "multi-stage-pump-sets": "Multi-Stage Pump & Pump Sets",
  "solar-pump-system": "Solar Pump System",
};

/* ════════════════════════════════════════════════════════
   SERIES DATA MAP
   Key: "product-slug/series-slug"
   ════════════════════════════════════════════════════════ */

export const seriesDataMap: Record<string, SeriesData> = {'''

L.append(header)

total_series = 0
total_models = 0

for cat_slug, series_list in all_data.items():
    feats = DEFAULT_FEATURES.get(cat_slug, DEFAULT_FEATURES["peripheral-pump"])
    apps  = DEFAULT_APPLICATIONS.get(cat_slug, DEFAULT_APPLICATIONS["peripheral-pump"])

    for series_key, series_data in series_list:
        key = f"{cat_slug}/{series_key}"
        title    = series_data["title"]    or series_key.replace("-", " ").upper()
        subtitle = series_data["subtitle"] or f"Professional {title} — Complete Range"
        models   = series_data["models"]
        total_series += 1
        total_models += len(models)

        img_path = get_series_image(series_key, cat_slug)

        # overview desc
        ov_desc = (
            f"The {title} delivers reliable performance across a wide range of flow and head conditions. "
            f"Built with quality materials and precision engineering, it is suitable for industrial, "
            f"commercial, and residential applications."
        )

        # model images: use series image (one per series)
        mi_label = title.replace("`", "").replace("\\", "")

        L.append(f'\n  /* ─── {title} ─── */')
        L.append(f'  "{key}": {{')
        L.append(f'    title: `{esc(title)}`,')
        L.append(f'    subtitle: `{esc(subtitle)}`,')
        L.append(f'    desc: `{esc(ov_desc)}`,')
        L.append(f'    heroCheckmarks: [')
        L.append(f'      "CE & ISO Certified",')
        L.append(f'      "Available in 50Hz/60Hz",')
        L.append(f'      "OEM/ODM Welcome",')
        L.append(f'    ],')
        L.append(f'    modelImages: [')
        L.append(f'      {{ src: "{img_path}", label: `{esc(mi_label)}` }},')
        L.append(f'    ],')
        L.append(f'    stats: [')
        L.append(f'      {{ icon: "📦", label: "Models", value: "{len(models)}+" }},')
        L.append(f'      {{ icon: "🏭", label: "CE / ISO", value: "Certified" }},')
        L.append(f'      {{ icon: "⚡", label: "50/60 Hz", value: "Available" }},')
        L.append(f'    ],')
        L.append(f'    overviewTitle: `{esc(title)} Overview`,')
        L.append(f'    overviewDesc: `{esc(ov_desc)}`,')
        L.append(f'    overviewImage: "{img_path}",')
        L.append(f'    phaseOptions: [')
        L.append(f'      {{ label: "Single Phase", desc: "1~ 220V 50Hz/60Hz" }},')
        L.append(f'      {{ label: "Three Phase", desc: "3~ 380V 50Hz/60Hz" }},')
        L.append(f'    ],')
        L.append(f'    tableCols: [')
        L.append(f'      {{ key: "model",       label: "Model" }},')
        L.append(f'      {{ key: "powerHP",     label: "Power (HP)" }},')
        L.append(f'      {{ key: "powerKW",     label: "Power (kW)" }},')
        L.append(f'      {{ key: "maxHead",     label: "Max Head (m)" }},')
        L.append(f'      {{ key: "maxFlow",     label: "Max Flow (m³/h)" }},')
        L.append(f'      {{ key: "suctionHead", label: "Suction Head (m)" }},')
        L.append(f'      {{ key: "inletOutlet", label: "Inlet/Outlet" }},')
        L.append(f'      {{ key: "weight",      label: "Weight (kg)" }},')
        L.append(f'    ],')
        L.append(f'    models: [')
        for m in models:
            L.append(f'      {{')
            L.append(f'        model: `{esc(m["model"])}`,')
            L.append(f'        powerHP: `{esc(m["powerHP"])}`,')
            L.append(f'        powerKW: `{esc(m["powerKW"])}`,')
            L.append(f'        maxHead: `{esc(m["maxHead"])}`,')
            L.append(f'        maxFlow: `{esc(m["maxFlow"])}`,')
            if m["suctionHead"]:
                L.append(f'        suctionHead: `{esc(m["suctionHead"])}`,')
            L.append(f'        inletOutlet: `{esc(m["inletOutlet"])}`,')
            L.append(f'        weight: `{esc(m["weight"])}`,')
            if m["maxParticle"]:
                L.append(f'        maxParticle: `{esc(m["maxParticle"])}`,')
            L.append(f'      }},')
        L.append(f'    ],')
        L.append(f'    performanceCurves: [],')
        L.append(f'    features: [')
        for icon, ftitle, fdesc in feats:
            L.append(f'      {{ icon: "{icon}", title: "{ftitle}", desc: "{fdesc}" }},')
        L.append(f'    ],')
        L.append(f'    applications: [')
        for icon, alabel, adesc in apps:
            L.append(f'      {{ icon: "{icon}", label: "{alabel}", desc: "{adesc}" }},')
        L.append(f'    ],')
        L.append(f'    faqs: [')
        for fq, fa in DEFAULT_FAQS:
            L.append(f'      {{ q: "{fq}", a: "{fa}" }},')
        L.append(f'    ],')
        L.append(f'    ctaTitle: `Ready to Order {esc(title)}?`,')
        L.append(f'    ctaDesc: `Contact our team for pricing, technical specs, and OEM/ODM options.`,')
        L.append(f'  }},')

L.append('\n};')
L.append('')

output = "\n".join(L)
with open(OUT_TS, "w", encoding="utf-8") as f:
    f.write(output)

print(f"\n=== DONE ===")
print(f"Total categories: {len(all_data)}")
print(f"Total series:     {total_series}")
print(f"Total models:     {total_models}")
print(f"Output written:   {OUT_TS}")
print(f"Lines:            {len(L)}")
